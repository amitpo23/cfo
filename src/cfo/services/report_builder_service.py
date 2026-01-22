"""
Report Builder & Scheduler Service
שירות בונה דוחות ותזמון
"""
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any, Callable
from dataclasses import dataclass, asdict, field
from enum import Enum
import json
import asyncio
from pathlib import Path
from sqlalchemy.orm import Session

from ..database import SessionLocal
from ..config import settings


class ReportFormat(str, Enum):
    """פורמט דוח"""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"
    HTML = "html"


class ReportFrequency(str, Enum):
    """תדירות דוח"""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    ON_DEMAND = "on_demand"


class ReportType(str, Enum):
    """סוג דוח"""
    PROFIT_LOSS = "profit_loss"
    BALANCE_SHEET = "balance_sheet"
    CASH_FLOW = "cash_flow"
    AGING_REPORT = "aging_report"
    BUDGET_VS_ACTUAL = "budget_vs_actual"
    KPI_DASHBOARD = "kpi_dashboard"
    VAT_REPORT = "vat_report"
    CUSTOM = "custom"


class DeliveryMethod(str, Enum):
    """שיטת משלוח"""
    EMAIL = "email"
    DOWNLOAD = "download"
    WEBHOOK = "webhook"
    SFTP = "sftp"
    GOOGLE_DRIVE = "google_drive"


@dataclass
class ReportColumn:
    """עמודת דוח"""
    field_name: str
    display_name: str
    data_type: str  # string, number, date, currency, percentage
    width: Optional[int] = None
    format_string: Optional[str] = None
    aggregation: Optional[str] = None  # sum, avg, count, min, max
    sortable: bool = True
    filterable: bool = True


@dataclass
class ReportFilter:
    """פילטר דוח"""
    field_name: str
    operator: str  # eq, ne, gt, gte, lt, lte, in, contains, between
    value: Any
    label: Optional[str] = None


@dataclass
class ReportTemplate:
    """תבנית דוח"""
    template_id: str
    name: str
    description: str
    report_type: ReportType
    columns: List[ReportColumn]
    default_filters: List[ReportFilter]
    grouping: List[str]
    sorting: List[Dict]  # [{'field': 'amount', 'direction': 'desc'}]
    summary_fields: List[str]
    created_by: str
    created_at: str
    is_public: bool
    organization_id: int


@dataclass
class ScheduledReport:
    """דוח מתוזמן"""
    schedule_id: str
    template_id: str
    name: str
    frequency: ReportFrequency
    next_run: str
    last_run: Optional[str]
    recipients: List[str]
    delivery_method: DeliveryMethod
    format: ReportFormat
    filters: List[ReportFilter]
    is_active: bool
    created_by: str
    organization_id: int
    parameters: Dict


@dataclass
class GeneratedReport:
    """דוח שנוצר"""
    report_id: str
    template_id: str
    generated_at: str
    format: ReportFormat
    file_path: Optional[str]
    file_size: int
    row_count: int
    generation_time_ms: int
    filters_applied: List[Dict]
    generated_by: str
    expires_at: str
    download_url: Optional[str]


@dataclass
class ReportExecution:
    """ביצוע דוח"""
    execution_id: str
    schedule_id: Optional[str]
    template_id: str
    status: str  # pending, running, completed, failed
    started_at: str
    completed_at: Optional[str]
    error_message: Optional[str]
    result: Optional[GeneratedReport]


class ReportBuilderService:
    """
    שירות בניית דוחות ותזמון
    Report Builder & Scheduler Service
    """
    
    def __init__(self, db: Session, organization_id: int = 1):
        self.db = db
        self.organization_id = organization_id
        self.reports_dir = Path(settings.reports_dir if hasattr(settings, 'reports_dir') else './reports')
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
        # תבניות ברירת מחדל
        self._default_templates = self._create_default_templates()
        
        # אחסון זמני (בפרודקשן - database)
        self._templates: Dict[str, ReportTemplate] = {}
        self._schedules: Dict[str, ScheduledReport] = {}
        self._executions: List[ReportExecution] = []
    
    # ===== Template Management =====
    
    def create_template(
        self,
        name: str,
        report_type: ReportType,
        columns: List[Dict],
        description: str = '',
        default_filters: Optional[List[Dict]] = None,
        grouping: Optional[List[str]] = None,
        sorting: Optional[List[Dict]] = None,
        summary_fields: Optional[List[str]] = None,
        is_public: bool = False,
        created_by: str = 'system'
    ) -> ReportTemplate:
        """
        יצירת תבנית דוח
        Create Report Template
        """
        import uuid
        
        template = ReportTemplate(
            template_id=f'TPL-{uuid.uuid4().hex[:8].upper()}',
            name=name,
            description=description,
            report_type=report_type,
            columns=[
                ReportColumn(**col) if isinstance(col, dict) else col
                for col in columns
            ],
            default_filters=[
                ReportFilter(**f) if isinstance(f, dict) else f
                for f in (default_filters or [])
            ],
            grouping=grouping or [],
            sorting=sorting or [],
            summary_fields=summary_fields or [],
            created_by=created_by,
            created_at=datetime.now().isoformat(),
            is_public=is_public,
            organization_id=self.organization_id
        )
        
        self._templates[template.template_id] = template
        return template
    
    def get_templates(
        self,
        report_type: Optional[ReportType] = None,
        include_public: bool = True
    ) -> List[ReportTemplate]:
        """
        קבלת תבניות
        Get Templates
        """
        templates = list(self._templates.values()) + list(self._default_templates.values())
        
        # סינון לפי סוג
        if report_type:
            templates = [t for t in templates if t.report_type == report_type]
        
        # סינון לפי ציבורי/פרטי
        if not include_public:
            templates = [t for t in templates if t.organization_id == self.organization_id]
        
        return templates
    
    def get_template(self, template_id: str) -> Optional[ReportTemplate]:
        """קבלת תבנית לפי ID"""
        return self._templates.get(template_id) or self._default_templates.get(template_id)
    
    def update_template(self, template_id: str, updates: Dict) -> Optional[ReportTemplate]:
        """עדכון תבנית"""
        template = self._templates.get(template_id)
        if not template:
            return None
        
        for key, value in updates.items():
            if hasattr(template, key):
                setattr(template, key, value)
        
        return template
    
    def delete_template(self, template_id: str) -> bool:
        """מחיקת תבנית"""
        if template_id in self._templates:
            del self._templates[template_id]
            return True
        return False
    
    # ===== Report Generation =====
    
    def generate_report(
        self,
        template_id: str,
        format: ReportFormat = ReportFormat.EXCEL,
        filters: Optional[List[Dict]] = None,
        parameters: Optional[Dict] = None,
        generated_by: str = 'system'
    ) -> GeneratedReport:
        """
        יצירת דוח
        Generate Report
        """
        import time
        import uuid
        
        start_time = time.time()
        
        template = self.get_template(template_id)
        if not template:
            raise ValueError(f"תבנית {template_id} לא נמצאה")
        
        # מיזוג פילטרים
        all_filters = list(template.default_filters)
        if filters:
            all_filters.extend([
                ReportFilter(**f) if isinstance(f, dict) else f
                for f in filters
            ])
        
        # הרצת הדוח לפי סוג
        data = self._execute_report_query(template, all_filters, parameters)
        
        # יצירת הקובץ
        report_id = f'RPT-{uuid.uuid4().hex[:8].upper()}'
        filename = f"{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format == ReportFormat.EXCEL:
            file_path = self._generate_excel(filename, template, data)
        elif format == ReportFormat.CSV:
            file_path = self._generate_csv(filename, template, data)
        elif format == ReportFormat.JSON:
            file_path = self._generate_json(filename, template, data)
        elif format == ReportFormat.HTML:
            file_path = self._generate_html(filename, template, data)
        else:
            file_path = self._generate_json(filename, template, data)
        
        generation_time = int((time.time() - start_time) * 1000)
        
        return GeneratedReport(
            report_id=report_id,
            template_id=template_id,
            generated_at=datetime.now().isoformat(),
            format=format,
            file_path=str(file_path) if file_path else None,
            file_size=file_path.stat().st_size if file_path and file_path.exists() else 0,
            row_count=len(data),
            generation_time_ms=generation_time,
            filters_applied=[asdict(f) if hasattr(f, '__dataclass_fields__') else f for f in all_filters],
            generated_by=generated_by,
            expires_at=(datetime.now() + timedelta(days=7)).isoformat(),
            download_url=f'/api/reports/download/{report_id}' if file_path else None
        )
    
    def preview_report(
        self,
        template_id: str,
        filters: Optional[List[Dict]] = None,
        parameters: Optional[Dict] = None,
        limit: int = 100
    ) -> Dict:
        """
        תצוגה מקדימה של דוח
        Preview Report
        """
        template = self.get_template(template_id)
        if not template:
            raise ValueError(f"תבנית {template_id} לא נמצאה")
        
        all_filters = list(template.default_filters)
        if filters:
            all_filters.extend([
                ReportFilter(**f) if isinstance(f, dict) else f
                for f in filters
            ])
        
        data = self._execute_report_query(template, all_filters, parameters)
        
        # חישוב סיכומים
        summary = {}
        for field in template.summary_fields:
            values = [row.get(field, 0) for row in data if isinstance(row.get(field), (int, float))]
            if values:
                summary[field] = {
                    'sum': sum(values),
                    'avg': sum(values) / len(values),
                    'min': min(values),
                    'max': max(values),
                    'count': len(values)
                }
        
        return {
            'template': asdict(template),
            'data': data[:limit],
            'total_rows': len(data),
            'summary': summary,
            'preview_limited': len(data) > limit
        }
    
    # ===== Scheduling =====
    
    def create_schedule(
        self,
        template_id: str,
        name: str,
        frequency: ReportFrequency,
        recipients: List[str],
        delivery_method: DeliveryMethod = DeliveryMethod.EMAIL,
        format: ReportFormat = ReportFormat.EXCEL,
        filters: Optional[List[Dict]] = None,
        parameters: Optional[Dict] = None,
        created_by: str = 'system'
    ) -> ScheduledReport:
        """
        יצירת תזמון דוח
        Create Report Schedule
        """
        import uuid
        
        schedule = ScheduledReport(
            schedule_id=f'SCH-{uuid.uuid4().hex[:8].upper()}',
            template_id=template_id,
            name=name,
            frequency=frequency,
            next_run=self._calculate_next_run(frequency),
            last_run=None,
            recipients=recipients,
            delivery_method=delivery_method,
            format=format,
            filters=[
                ReportFilter(**f) if isinstance(f, dict) else f
                for f in (filters or [])
            ],
            is_active=True,
            created_by=created_by,
            organization_id=self.organization_id,
            parameters=parameters or {}
        )
        
        self._schedules[schedule.schedule_id] = schedule
        return schedule
    
    def get_schedules(
        self,
        template_id: Optional[str] = None,
        active_only: bool = False
    ) -> List[ScheduledReport]:
        """
        קבלת תזמונים
        Get Schedules
        """
        schedules = list(self._schedules.values())
        
        if template_id:
            schedules = [s for s in schedules if s.template_id == template_id]
        
        if active_only:
            schedules = [s for s in schedules if s.is_active]
        
        return schedules
    
    def update_schedule(self, schedule_id: str, updates: Dict) -> Optional[ScheduledReport]:
        """עדכון תזמון"""
        schedule = self._schedules.get(schedule_id)
        if not schedule:
            return None
        
        for key, value in updates.items():
            if hasattr(schedule, key):
                if key == 'frequency':
                    setattr(schedule, key, ReportFrequency(value))
                    schedule.next_run = self._calculate_next_run(schedule.frequency)
                else:
                    setattr(schedule, key, value)
        
        return schedule
    
    def delete_schedule(self, schedule_id: str) -> bool:
        """מחיקת תזמון"""
        if schedule_id in self._schedules:
            del self._schedules[schedule_id]
            return True
        return False
    
    def pause_schedule(self, schedule_id: str) -> bool:
        """השהיית תזמון"""
        schedule = self._schedules.get(schedule_id)
        if schedule:
            schedule.is_active = False
            return True
        return False
    
    def resume_schedule(self, schedule_id: str) -> bool:
        """חידוש תזמון"""
        schedule = self._schedules.get(schedule_id)
        if schedule:
            schedule.is_active = True
            schedule.next_run = self._calculate_next_run(schedule.frequency)
            return True
        return False
    
    # ===== Execution =====
    
    async def run_scheduled_reports(self) -> List[ReportExecution]:
        """
        הרצת דוחות מתוזמנים
        Run Scheduled Reports
        """
        import uuid
        
        now = datetime.now()
        executions = []
        
        for schedule in self._schedules.values():
            if not schedule.is_active:
                continue
            
            next_run = datetime.fromisoformat(schedule.next_run)
            if next_run <= now:
                execution = ReportExecution(
                    execution_id=f'EXC-{uuid.uuid4().hex[:8].upper()}',
                    schedule_id=schedule.schedule_id,
                    template_id=schedule.template_id,
                    status='running',
                    started_at=datetime.now().isoformat(),
                    completed_at=None,
                    error_message=None,
                    result=None
                )
                
                try:
                    # יצירת הדוח
                    report = self.generate_report(
                        template_id=schedule.template_id,
                        format=schedule.format,
                        filters=[asdict(f) for f in schedule.filters],
                        parameters=schedule.parameters,
                        generated_by=f'schedule:{schedule.schedule_id}'
                    )
                    
                    # משלוח
                    await self._deliver_report(schedule, report)
                    
                    execution.status = 'completed'
                    execution.result = report
                    
                    # עדכון תזמון
                    schedule.last_run = datetime.now().isoformat()
                    schedule.next_run = self._calculate_next_run(schedule.frequency)
                    
                except Exception as e:
                    execution.status = 'failed'
                    execution.error_message = str(e)
                
                execution.completed_at = datetime.now().isoformat()
                self._executions.append(execution)
                executions.append(execution)
        
        return executions
    
    def get_execution_history(
        self,
        schedule_id: Optional[str] = None,
        template_id: Optional[str] = None,
        limit: int = 50
    ) -> List[ReportExecution]:
        """
        היסטוריית ביצועים
        Execution History
        """
        executions = self._executions
        
        if schedule_id:
            executions = [e for e in executions if e.schedule_id == schedule_id]
        
        if template_id:
            executions = [e for e in executions if e.template_id == template_id]
        
        # מיון לפי תאריך (חדש ראשון)
        executions = sorted(executions, key=lambda x: x.started_at, reverse=True)
        
        return executions[:limit]
    
    # ===== Private Methods =====
    
    def _create_default_templates(self) -> Dict[str, ReportTemplate]:
        """יצירת תבניות ברירת מחדל"""
        templates = {}
        
        # תבנית רווח והפסד
        templates['DEFAULT-PL'] = ReportTemplate(
            template_id='DEFAULT-PL',
            name='דוח רווח והפסד',
            description='דוח רווח והפסד חודשי/שנתי',
            report_type=ReportType.PROFIT_LOSS,
            columns=[
                ReportColumn('category', 'קטגוריה', 'string', 200),
                ReportColumn('amount', 'סכום', 'currency', 120, '₪#,##0'),
                ReportColumn('budget', 'תקציב', 'currency', 120, '₪#,##0'),
                ReportColumn('variance', 'סטייה', 'percentage', 100, '0.0%'),
                ReportColumn('previous_period', 'תקופה קודמת', 'currency', 120, '₪#,##0'),
            ],
            default_filters=[],
            grouping=['category'],
            sorting=[{'field': 'amount', 'direction': 'desc'}],
            summary_fields=['amount', 'budget'],
            created_by='system',
            created_at=datetime.now().isoformat(),
            is_public=True,
            organization_id=0
        )
        
        # תבנית גיול חובות
        templates['DEFAULT-AGING'] = ReportTemplate(
            template_id='DEFAULT-AGING',
            name='דוח גיול חובות',
            description='דוח גיול חובות לקוחות',
            report_type=ReportType.AGING_REPORT,
            columns=[
                ReportColumn('customer_name', 'לקוח', 'string', 200),
                ReportColumn('current', 'שוטף', 'currency', 100, '₪#,##0'),
                ReportColumn('days_31_60', '31-60 יום', 'currency', 100, '₪#,##0'),
                ReportColumn('days_61_90', '61-90 יום', 'currency', 100, '₪#,##0'),
                ReportColumn('days_91_120', '91-120 יום', 'currency', 100, '₪#,##0'),
                ReportColumn('over_120', 'מעל 120', 'currency', 100, '₪#,##0'),
                ReportColumn('total', 'סה"כ', 'currency', 120, '₪#,##0'),
            ],
            default_filters=[],
            grouping=[],
            sorting=[{'field': 'total', 'direction': 'desc'}],
            summary_fields=['current', 'days_31_60', 'days_61_90', 'days_91_120', 'over_120', 'total'],
            created_by='system',
            created_at=datetime.now().isoformat(),
            is_public=True,
            organization_id=0
        )
        
        # תבנית KPI
        templates['DEFAULT-KPI'] = ReportTemplate(
            template_id='DEFAULT-KPI',
            name='דשבורד KPI',
            description='סיכום מדדי ביצוע מרכזיים',
            report_type=ReportType.KPI_DASHBOARD,
            columns=[
                ReportColumn('kpi_name', 'מדד', 'string', 200),
                ReportColumn('value', 'ערך', 'number', 100),
                ReportColumn('target', 'יעד', 'number', 100),
                ReportColumn('status', 'סטטוס', 'string', 80),
                ReportColumn('trend', 'מגמה', 'string', 80),
            ],
            default_filters=[],
            grouping=['category'],
            sorting=[{'field': 'category', 'direction': 'asc'}],
            summary_fields=[],
            created_by='system',
            created_at=datetime.now().isoformat(),
            is_public=True,
            organization_id=0
        )
        
        return templates
    
    def _execute_report_query(
        self,
        template: ReportTemplate,
        filters: List[ReportFilter],
        parameters: Optional[Dict]
    ) -> List[Dict]:
        """ביצוע שאילתת הדוח"""
        import random
        
        # בפרודקשן - שאילתה אמיתית לDB
        # כאן - נתוני דוגמה
        
        if template.report_type == ReportType.PROFIT_LOSS:
            return self._generate_pl_data()
        elif template.report_type == ReportType.AGING_REPORT:
            return self._generate_aging_data()
        elif template.report_type == ReportType.KPI_DASHBOARD:
            return self._generate_kpi_data()
        elif template.report_type == ReportType.BUDGET_VS_ACTUAL:
            return self._generate_budget_data()
        else:
            return []
    
    def _generate_pl_data(self) -> List[Dict]:
        """יצירת נתוני רווח והפסד"""
        import random
        
        categories = [
            ('הכנסות', 'הכנסות ממכירות', 500000, 480000),
            ('הכנסות', 'הכנסות משירותים', 150000, 120000),
            ('עלות המכר', 'חומרים', -180000, -160000),
            ('עלות המכר', 'עבודה ישירה', -80000, -75000),
            ('הוצאות תפעול', 'שכר', -120000, -115000),
            ('הוצאות תפעול', 'שיווק', -35000, -40000),
            ('הוצאות תפעול', 'משרד', -25000, -22000),
            ('הוצאות מימון', 'ריבית', -8000, -10000),
        ]
        
        return [
            {
                'category': cat,
                'subcategory': subcat,
                'amount': amount + random.randint(-5000, 5000),
                'budget': budget,
                'variance': ((amount - budget) / abs(budget) * 100) if budget else 0,
                'previous_period': amount * 0.95
            }
            for cat, subcat, amount, budget in categories
        ]
    
    def _generate_aging_data(self) -> List[Dict]:
        """יצירת נתוני גיול"""
        import random
        
        customers = ['לקוח א', 'לקוח ב', 'לקוח ג', 'לקוח ד', 'לקוח ה']
        
        data = []
        for customer in customers:
            current = random.randint(0, 50000)
            d31_60 = random.randint(0, 30000)
            d61_90 = random.randint(0, 20000)
            d91_120 = random.randint(0, 15000)
            over_120 = random.randint(0, 10000)
            
            data.append({
                'customer_name': customer,
                'current': current,
                'days_31_60': d31_60,
                'days_61_90': d61_90,
                'days_91_120': d91_120,
                'over_120': over_120,
                'total': current + d31_60 + d61_90 + d91_120 + over_120
            })
        
        return data
    
    def _generate_kpi_data(self) -> List[Dict]:
        """יצירת נתוני KPI"""
        kpis = [
            ('רווחיות', 'שיעור רווח גולמי', 42.5, 40, 'above_target', 'up'),
            ('רווחיות', 'שיעור רווח נקי', 15.2, 12, 'above_target', 'up'),
            ('נזילות', 'יחס שוטף', 1.85, 1.5, 'above_target', 'stable'),
            ('יעילות', 'DSO', 45, 35, 'below_target', 'down'),
            ('צמיחה', 'צמיחת הכנסות', 8.5, 10, 'below_target', 'up'),
        ]
        
        return [
            {
                'category': cat,
                'kpi_name': name,
                'value': value,
                'target': target,
                'status': status,
                'trend': trend
            }
            for cat, name, value, target, status, trend in kpis
        ]
    
    def _generate_budget_data(self) -> List[Dict]:
        """יצירת נתוני תקציב"""
        import random
        
        categories = ['הכנסות', 'שכר', 'שיווק', 'תפעול', 'הנה"כ']
        
        return [
            {
                'category': cat,
                'budget': budget,
                'actual': budget + random.randint(-int(budget * 0.2), int(budget * 0.2)),
                'variance': 0,
                'variance_pct': 0
            }
            for cat, budget in [
                ('הכנסות', 500000),
                ('שכר', -150000),
                ('שיווק', -50000),
                ('תפעול', -80000),
                ('הנה"כ', -30000),
            ]
        ]
    
    def _generate_excel(self, filename: str, template: ReportTemplate, data: List[Dict]) -> Path:
        """יצירת קובץ Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return self._generate_csv(filename, template, data)
        
        wb = Workbook()
        ws = wb.active
        ws.title = template.name[:31]  # Excel limit
        
        # RTL
        ws.sheet_view.rightToLeft = True
        
        # כותרת
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(template.columns))
        ws['A1'] = template.name
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # תאריך יצירה
        ws['A2'] = f'תאריך הפקה: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # כותרות עמודות
        header_row = 4
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_idx, col in enumerate(template.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col.display_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            if col.width:
                ws.column_dimensions[get_column_letter(col_idx)].width = col.width / 7
        
        # נתונים
        for row_idx, row_data in enumerate(data, header_row + 1):
            for col_idx, col in enumerate(template.columns, 1):
                value = row_data.get(col.field_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col.data_type == 'currency':
                    cell.number_format = '₪#,##0'
                elif col.data_type == 'percentage':
                    cell.number_format = '0.0%'
                    if isinstance(value, (int, float)):
                        cell.value = value / 100
        
        # שמירה
        file_path = self.reports_dir / f'{filename}.xlsx'
        wb.save(file_path)
        
        return file_path
    
    def _generate_csv(self, filename: str, template: ReportTemplate, data: List[Dict]) -> Path:
        """יצירת קובץ CSV"""
        import csv
        
        file_path = self.reports_dir / f'{filename}.csv'
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # כותרות
            headers = [col.display_name for col in template.columns]
            writer.writerow(headers)
            
            # נתונים
            for row in data:
                row_values = [row.get(col.field_name, '') for col in template.columns]
                writer.writerow(row_values)
        
        return file_path
    
    def _generate_json(self, filename: str, template: ReportTemplate, data: List[Dict]) -> Path:
        """יצירת קובץ JSON"""
        file_path = self.reports_dir / f'{filename}.json'
        
        output = {
            'template': asdict(template),
            'generated_at': datetime.now().isoformat(),
            'data': data,
            'row_count': len(data)
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        
        return file_path
    
    def _generate_html(self, filename: str, template: ReportTemplate, data: List[Dict]) -> Path:
        """יצירת קובץ HTML"""
        file_path = self.reports_dir / f'{filename}.html'
        
        html = f"""<!DOCTYPE html>
<html dir="rtl" lang="he">
<head>
    <meta charset="UTF-8">
    <title>{template.name}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; direction: rtl; }}
        h1 {{ color: #366092; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th {{ background-color: #366092; color: white; padding: 12px; text-align: right; }}
        td {{ border: 1px solid #ddd; padding: 10px; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .meta {{ color: #666; margin-bottom: 20px; }}
        .currency {{ text-align: left; }}
        .number {{ text-align: left; }}
    </style>
</head>
<body>
    <h1>{template.name}</h1>
    <p class="meta">תאריך הפקה: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
    <table>
        <thead>
            <tr>
                {''.join(f'<th>{col.display_name}</th>' for col in template.columns)}
            </tr>
        </thead>
        <tbody>
"""
        
        for row in data:
            html += '<tr>'
            for col in template.columns:
                value = row.get(col.field_name, '')
                if col.data_type == 'currency' and isinstance(value, (int, float)):
                    value = f'₪{value:,.0f}'
                elif col.data_type == 'percentage' and isinstance(value, (int, float)):
                    value = f'{value:.1f}%'
                html += f'<td class="{col.data_type}">{value}</td>'
            html += '</tr>\n'
        
        html += """
        </tbody>
    </table>
</body>
</html>
"""
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html)
        
        return file_path
    
    def _calculate_next_run(self, frequency: ReportFrequency) -> str:
        """חישוב הרצה הבאה"""
        now = datetime.now()
        
        if frequency == ReportFrequency.DAILY:
            next_run = now + timedelta(days=1)
            next_run = next_run.replace(hour=6, minute=0, second=0, microsecond=0)
        elif frequency == ReportFrequency.WEEKLY:
            days_until_sunday = (6 - now.weekday()) % 7 or 7
            next_run = now + timedelta(days=days_until_sunday)
            next_run = next_run.replace(hour=6, minute=0, second=0, microsecond=0)
        elif frequency == ReportFrequency.MONTHLY:
            if now.month == 12:
                next_run = now.replace(year=now.year + 1, month=1, day=1)
            else:
                next_run = now.replace(month=now.month + 1, day=1)
            next_run = next_run.replace(hour=6, minute=0, second=0, microsecond=0)
        elif frequency == ReportFrequency.QUARTERLY:
            current_quarter = (now.month - 1) // 3 + 1
            next_quarter_month = ((current_quarter % 4) * 3) + 1
            if next_quarter_month <= now.month:
                next_run = now.replace(year=now.year + 1, month=next_quarter_month, day=1)
            else:
                next_run = now.replace(month=next_quarter_month, day=1)
            next_run = next_run.replace(hour=6, minute=0, second=0, microsecond=0)
        elif frequency == ReportFrequency.YEARLY:
            next_run = now.replace(year=now.year + 1, month=1, day=1, hour=6, minute=0, second=0, microsecond=0)
        else:
            next_run = now
        
        return next_run.isoformat()
    
    async def _deliver_report(self, schedule: ScheduledReport, report: GeneratedReport):
        """משלוח הדוח"""
        if schedule.delivery_method == DeliveryMethod.EMAIL:
            await self._send_email(schedule.recipients, report)
        elif schedule.delivery_method == DeliveryMethod.WEBHOOK:
            await self._send_webhook(schedule.parameters.get('webhook_url'), report)
        # שאר שיטות המשלוח...
    
    async def _send_email(self, recipients: List[str], report: GeneratedReport):
        """שליחת מייל"""
        # בפרודקשן - שילוב עם שירות מייל
        pass
    
    async def _send_webhook(self, url: str, report: GeneratedReport):
        """שליחת webhook"""
        # בפרודקשן - HTTP POST
        pass
