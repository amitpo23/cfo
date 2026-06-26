"""
Budget Management Service
שירות ניהול תקציב ובקרה תקציבית
"""
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict, field
from enum import Enum
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract

from ..models import Transaction, Account, Budget
from ..database import SessionLocal

logger = logging.getLogger(__name__)


class BudgetPeriod(str, Enum):
    """תקופת תקציב"""
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"


class BudgetStatus(str, Enum):
    """סטטוס תקציב"""
    ON_TRACK = "on_track"
    WARNING = "warning"
    OVER_BUDGET = "over_budget"
    UNDER_BUDGET = "under_budget"


@dataclass
class BudgetCategory:
    """קטגוריית תקציב"""
    category_id: str
    category_name: str
    category_hebrew: str
    budget_amount: float
    actual_amount: float
    variance: float
    variance_percentage: float
    status: BudgetStatus
    forecast_end_of_period: float
    monthly_breakdown: List[Dict] = field(default_factory=list)


@dataclass
class BudgetSummary:
    """סיכום תקציב"""
    period: str
    period_start: str
    period_end: str
    total_budget: float
    total_actual: float
    total_variance: float
    variance_percentage: float
    status: BudgetStatus
    categories: List[BudgetCategory]
    alerts: List[Dict]


@dataclass
class BudgetAlert:
    """התראת תקציב"""
    alert_id: str
    category: str
    category_hebrew: str
    alert_type: str
    message: str
    severity: str  # low, medium, high, critical
    created_at: str
    budget_amount: float
    actual_amount: float
    variance_percentage: float


@dataclass
class ScenarioAnalysis:
    """ניתוח תרחישים"""
    scenario_name: str
    scenario_type: str  # best, worst, expected
    total_revenue: float
    total_expenses: float
    net_income: float
    cash_flow: float
    assumptions: Dict
    monthly_projections: List[Dict]


# מיפוי קטגוריות לעברית
BUDGET_CATEGORIES_HEBREW = {
    'revenue': 'הכנסות',
    'sales': 'מכירות',
    'services': 'שירותים',
    'salary': 'משכורות',
    'rent': 'שכירות',
    'utilities': 'חשמל/מים/גז',
    'marketing': 'שיווק ופרסום',
    'office': 'הוצאות משרד',
    'travel': 'נסיעות',
    'insurance': 'ביטוח',
    'professional': 'שירותים מקצועיים',
    'software': 'תוכנה ומחשוב',
    'equipment': 'ציוד',
    'maintenance': 'אחזקה',
    'taxes': 'מסים ואגרות',
    'finance': 'הוצאות מימון',
    'other_income': 'הכנסות אחרות',
    'other_expense': 'הוצאות אחרות'
}


class BudgetService:
    """
    שירות ניהול תקציב
    Budget Management Service
    """
    
    def __init__(self, db: Session, organization_id: int = 1):
        self.db = db
        self.organization_id = organization_id
        # אחסון תקציבים (בפרודקשן יהיה בDB)
        self._budgets: Dict[str, Dict] = {}
        self._alerts: List[BudgetAlert] = []
    
    def create_budget(
        self,
        category: str,
        planned_amount: float,
        period_start: date,
        period_end: Optional[date] = None,
        notes: Optional[str] = None,
    ) -> Dict:
        """יצירת/עדכון תקציב לקטגוריה (נשמר ב-DB, לפי שנה+חודש של תחילת התקופה)."""
        year = period_start.year
        month = period_start.month
        row = (
            self.db.query(Budget)
            .filter(
                Budget.organization_id == self.organization_id,
                Budget.category_name == category,
                Budget.year == year,
                Budget.month == month,
            )
            .first()
        )
        if row is None:
            row = Budget(
                organization_id=self.organization_id,
                category_name=category,
                year=year,
                month=month,
            )
            self.db.add(row)
        row.budgeted_amount = planned_amount
        row.notes = notes
        self.db.commit()
        self.db.refresh(row)
        return {
            "id": row.id,
            "category": category,
            "year": year,
            "month": month,
            "budgeted_amount": float(row.budgeted_amount or 0),
        }

    def bulk_upsert(self, items: List[Dict]) -> Dict:
        """הזנת תקציבים מרובים בבת אחת. כל פריט: {category, year, month, amount}."""
        count = 0
        for it in items:
            category = it.get("category")
            amount = it.get("amount")
            year = it.get("year")
            month = it.get("month")
            if not category or year is None or month is None or amount is None:
                continue
            self.create_budget(
                category=category,
                planned_amount=float(amount),
                period_start=date(int(year), int(month), 1),
            )
            count += 1
        return {"saved": count}

    def import_from_excel(self, content: bytes, default_year: Optional[int] = None) -> Dict:
        """ייבוא תקציבים מקובץ Excel.

        פורמט נתמך — שורת כותרת ואז שורות:
        עמודה A=קטגוריה, B=שנה, C=חודש, D=סכום.
        אם אין שנה/חודש בשורה — משתמשים ב-default_year והחודש הנוכחי.
        """
        try:
            import openpyxl
            import io
        except ImportError:
            raise ValueError("openpyxl required for Excel import")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        items: List[Dict] = []
        skipped = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # דילוג על שורת כותרת
            if not row or row[0] is None:
                continue
            category = str(row[0]).strip()
            year = row[1] if len(row) > 1 and row[1] else default_year
            month = row[2] if len(row) > 2 and row[2] else None
            amount = row[3] if len(row) > 3 else None
            try:
                if not category or year is None or month is None or amount is None:
                    skipped += 1
                    continue
                items.append({
                    "category": category,
                    "year": int(year),
                    "month": int(month),
                    "amount": float(amount),
                })
            except (TypeError, ValueError):
                skipped += 1
        result = self.bulk_upsert(items)
        result["skipped"] = skipped
        return result

    def export_template(self) -> bytes:
        """תבנית Excel להזנת תקציב."""
        import openpyxl
        import io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "תקציב"
        ws.sheet_view.rightToLeft = True
        ws.append(["קטגוריה", "שנה", "חודש", "סכום מתוכנן"])
        today = date.today()
        for cat in ("sales", "materials", "rent", "salaries", "marketing"):
            ws.append([cat, today.year, today.month, 0])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _load_budgets(self, year: int, month: Optional[int]) -> Dict[str, float]:
        """טעינת תקציבים מה-DB לתקופה. ריק -> ברירת מחדל."""
        q = self.db.query(Budget).filter(
            Budget.organization_id == self.organization_id,
            Budget.year == year,
        )
        if month:
            q = q.filter(Budget.month == month)
        budgets: Dict[str, float] = {}
        for row in q.all():
            key = row.category_name or (str(row.category_id) if row.category_id else "כללי")
            budgets[key] = budgets.get(key, 0) + float(row.budgeted_amount or 0)
        # ללא fallback לתקציב-דמה: תקציב שלא הוגדר מחזיר ריק (אפסים כנים),
        # כדי לא להציג יעדים מומצאים שהמשתמש לא הזין.
        return budgets
    
    def get_budget_vs_actual(
        self,
        year: int,
        month: Optional[int] = None,
        period: BudgetPeriod = BudgetPeriod.MONTHLY
    ) -> BudgetSummary:
        """
        השוואת תקציב מול ביצוע
        Budget vs Actual comparison
        """
        # קביעת טווח תאריכים
        if period == BudgetPeriod.MONTHLY and month:
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1)
            else:
                end_date = date(year, month + 1, 1)
            period_name = f"{year}-{month:02d}"
        elif period == BudgetPeriod.QUARTERLY and month:
            quarter = (month - 1) // 3 + 1
            start_month = (quarter - 1) * 3 + 1
            start_date = date(year, start_month, 1)
            end_month = start_month + 3
            if end_month > 12:
                end_date = date(year + 1, 1, 1)
            else:
                end_date = date(year, end_month, 1)
            period_name = f"Q{quarter}/{year}"
        else:
            start_date = date(year, 1, 1)
            end_date = date(year + 1, 1, 1)
            period_name = str(year)
        
        # שליפת נתוני ביצוע מהDB
        actuals = self._get_actual_by_category(start_date, end_date)
        
        # טעינת תקציב מה-DB
        category_budgets = self._load_budgets(year, month)
        
        # חישוב לפי קטגוריה
        categories = []
        total_budget = 0
        total_actual = 0
        alerts = []
        
        for cat_id, budget_amount in category_budgets.items():
            actual_amount = actuals.get(cat_id, 0)
            variance = budget_amount - actual_amount
            variance_pct = (variance / budget_amount * 100) if budget_amount else 0
            
            # קביעת סטטוס
            if cat_id.startswith('revenue') or cat_id in ['sales', 'services', 'other_income']:
                # הכנסות - רוצים יותר
                if actual_amount >= budget_amount:
                    status = BudgetStatus.ON_TRACK
                elif actual_amount >= budget_amount * 0.8:
                    status = BudgetStatus.WARNING
                else:
                    status = BudgetStatus.UNDER_BUDGET
            else:
                # הוצאות - רוצים פחות
                if actual_amount <= budget_amount:
                    status = BudgetStatus.ON_TRACK
                elif actual_amount <= budget_amount * 1.1:
                    status = BudgetStatus.WARNING
                else:
                    status = BudgetStatus.OVER_BUDGET
            
            # חיזוי סוף תקופה
            days_passed = (datetime.now().date() - start_date).days
            total_days = (end_date - start_date).days
            if days_passed > 0 and total_days > 0:
                daily_rate = actual_amount / days_passed
                forecast = daily_rate * total_days
            else:
                forecast = actual_amount
            
            category = BudgetCategory(
                category_id=cat_id,
                category_name=cat_id,
                category_hebrew=BUDGET_CATEGORIES_HEBREW.get(cat_id, cat_id),
                budget_amount=budget_amount,
                actual_amount=actual_amount,
                variance=variance,
                variance_percentage=variance_pct,
                status=status,
                forecast_end_of_period=forecast
            )
            categories.append(category)
            
            total_budget += budget_amount
            total_actual += actual_amount
            
            # יצירת התראות
            if status == BudgetStatus.OVER_BUDGET:
                alerts.append({
                    'category': cat_id,
                    'category_hebrew': BUDGET_CATEGORIES_HEBREW.get(cat_id, cat_id),
                    'type': 'over_budget',
                    'message': f"חריגה מתקציב ב{abs(variance_pct):.1f}%",
                    'severity': 'high' if abs(variance_pct) > 20 else 'medium'
                })
            elif status == BudgetStatus.UNDER_BUDGET and cat_id in ['sales', 'revenue', 'services']:
                alerts.append({
                    'category': cat_id,
                    'category_hebrew': BUDGET_CATEGORIES_HEBREW.get(cat_id, cat_id),
                    'type': 'under_target',
                    'message': f"הכנסות נמוכות מהיעד ב{abs(variance_pct):.1f}%",
                    'severity': 'high' if abs(variance_pct) > 20 else 'medium'
                })
        
        # סטטוס כללי
        total_variance = total_budget - total_actual
        total_variance_pct = (total_variance / total_budget * 100) if total_budget else 0
        
        if abs(total_variance_pct) <= 5:
            overall_status = BudgetStatus.ON_TRACK
        elif total_variance_pct > 5:
            overall_status = BudgetStatus.UNDER_BUDGET
        else:
            overall_status = BudgetStatus.OVER_BUDGET
        
        return BudgetSummary(
            period=period_name,
            period_start=start_date.isoformat(),
            period_end=end_date.isoformat(),
            total_budget=total_budget,
            total_actual=total_actual,
            total_variance=total_variance,
            variance_percentage=total_variance_pct,
            status=overall_status,
            categories=categories,
            alerts=alerts
        )
    
    def get_budget_alerts(
        self,
        threshold_percentage: float = 10.0
    ) -> List[BudgetAlert]:
        """
        קבלת התראות תקציב
        Get budget alerts
        """
        now = datetime.now()
        summary = self.get_budget_vs_actual(
            year=now.year,
            month=now.month
        )
        
        alerts = []
        for cat in summary.categories:
            if abs(cat.variance_percentage) >= threshold_percentage:
                severity = 'critical' if abs(cat.variance_percentage) >= 30 else \
                          'high' if abs(cat.variance_percentage) >= 20 else \
                          'medium' if abs(cat.variance_percentage) >= 10 else 'low'
                
                alert = BudgetAlert(
                    alert_id=f"{cat.category_id}_{now.strftime('%Y%m')}",
                    category=cat.category_id,
                    category_hebrew=cat.category_hebrew,
                    alert_type='over_budget' if cat.status == BudgetStatus.OVER_BUDGET else 'variance',
                    message=self._generate_alert_message(cat),
                    severity=severity,
                    created_at=now.isoformat(),
                    budget_amount=cat.budget_amount,
                    actual_amount=cat.actual_amount,
                    variance_percentage=cat.variance_percentage
                )
                alerts.append(alert)
        
        return sorted(alerts, key=lambda x: abs(x.variance_percentage), reverse=True)
    
    def run_scenario_analysis(
        self,
        base_year: int,
        scenarios: List[str] = ['best', 'worst', 'expected']
    ) -> List[ScenarioAnalysis]:
        """
        ניתוח תרחישים
        Scenario Analysis
        """
        results = []
        
        # נתונים בסיסיים
        base_summary = self.get_budget_vs_actual(year=base_year)
        
        for scenario in scenarios:
            if scenario == 'best':
                revenue_factor = 1.2
                expense_factor = 0.9
                assumptions = {
                    'revenue_growth': '+20%',
                    'expense_reduction': '-10%',
                    'description': 'תרחיש אופטימי - גידול במכירות וצמצום הוצאות'
                }
            elif scenario == 'worst':
                revenue_factor = 0.8
                expense_factor = 1.15
                assumptions = {
                    'revenue_growth': '-20%',
                    'expense_increase': '+15%',
                    'description': 'תרחיש פסימי - ירידה במכירות ועלייה בהוצאות'
                }
            else:  # expected
                revenue_factor = 1.05
                expense_factor = 1.03
                assumptions = {
                    'revenue_growth': '+5%',
                    'expense_increase': '+3%',
                    'description': 'תרחיש צפוי - צמיחה מתונה'
                }
            
            # חישוב
            revenue_cats = [c for c in base_summary.categories 
                          if c.category_id in ['sales', 'revenue', 'services', 'other_income']]
            expense_cats = [c for c in base_summary.categories 
                          if c.category_id not in ['sales', 'revenue', 'services', 'other_income']]
            
            total_revenue = sum(c.actual_amount for c in revenue_cats) * revenue_factor
            total_expenses = sum(c.actual_amount for c in expense_cats) * expense_factor
            net_income = total_revenue - total_expenses
            
            # תחזית חודשית
            monthly = []
            for month in range(1, 13):
                monthly.append({
                    'month': f"{base_year}-{month:02d}",
                    'revenue': total_revenue / 12,
                    'expenses': total_expenses / 12,
                    'net': net_income / 12
                })
            
            results.append(ScenarioAnalysis(
                scenario_name=f"תרחיש {scenario}",
                scenario_type=scenario,
                total_revenue=total_revenue,
                total_expenses=total_expenses,
                net_income=net_income,
                cash_flow=net_income * 0.8,  # הנחה גסה
                assumptions=assumptions,
                monthly_projections=monthly
            ))

        return results

    def project_scenario(
        self,
        revenue_change_pct: float = 0,
        expense_change_pct: float = 0,
        base_year: Optional[int] = None,
    ) -> Dict:
        """תרחיש מותאם: השפעת שינוי באחוזי הכנסות/הוצאות על השורה התחתונה."""
        from datetime import date as _date
        base_summary = self.get_budget_vs_actual(year=base_year or _date.today().year)
        revenue_ids = {'sales', 'revenue', 'services', 'other_income'}
        base_revenue = sum(
            c.actual_amount for c in base_summary.categories
            if c.category_id in revenue_ids
        )
        base_expenses = sum(
            c.actual_amount for c in base_summary.categories
            if c.category_id not in revenue_ids
        )
        projected_revenue = base_revenue * (1 + revenue_change_pct / 100)
        projected_expenses = base_expenses * (1 + expense_change_pct / 100)
        return {
            'base_revenue': base_revenue,
            'base_expenses': base_expenses,
            'projected_revenue': projected_revenue,
            'projected_expenses': projected_expenses,
            'projected_net_income': projected_revenue - projected_expenses,
            'revenue_change_pct': revenue_change_pct,
            'expense_change_pct': expense_change_pct,
        }

    def get_budget_trend(
        self,
        category_id: str,
        months: int = 12
    ) -> List[Dict]:
        """
        מגמת תקציב לאורך זמן
        Budget trend over time
        """
        now = datetime.now()
        trend = []
        
        for i in range(months - 1, -1, -1):
            month = now.month - i
            year = now.year
            while month <= 0:
                month += 12
                year -= 1
            
            summary = self.get_budget_vs_actual(year=year, month=month)
            cat = next((c for c in summary.categories if c.category_id == category_id), None)
            
            if cat:
                trend.append({
                    'period': f"{year}-{month:02d}",
                    'budget': cat.budget_amount,
                    'actual': cat.actual_amount,
                    'variance': cat.variance,
                    'variance_percentage': cat.variance_percentage
                })
        
        return trend
    
    def _get_actual_by_category(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, float]:
        """שליפת ביצוע בפועל לפי קטגוריה — נתונים אמיתיים בלבד.

        ללא fallback אקראי שקט: בכשל מחזיר {} ורושם לוג, כדי שלא יוצגו
        למשתמש מספרים מזויפים כאילו הם הביצוע בפועל.
        """
        try:
            transactions = self.db.query(Transaction).filter(
                Transaction.organization_id == self.organization_id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date < end_date,
            ).all()
        except Exception:
            logger.exception("budget actuals query failed for org %s", self.organization_id)
            return {}

        actuals: Dict[str, float] = {}
        for tx in transactions:
            category = self._categorize_transaction(tx)
            actuals[category] = actuals.get(category, 0) + float(tx.amount)
        return actuals
    
    def _categorize_transaction(self, tx: Transaction) -> str:
        """קטגוריזציה של עסקה"""
        description = (tx.description or '').lower()
        
        # מיפוי מילות מפתח לקטגוריות
        keywords = {
            'salary': ['משכורת', 'שכר', 'salary'],
            'rent': ['שכירות', 'rent'],
            'utilities': ['חשמל', 'מים', 'גז', 'ארנונה'],
            'marketing': ['פרסום', 'שיווק', 'marketing', 'google', 'facebook'],
            'software': ['תוכנה', 'software', 'saas', 'subscription'],
            'insurance': ['ביטוח', 'insurance'],
            'professional': ['עו"ד', 'רו"ח', 'יועץ', 'consultant'],
            'office': ['משרד', 'office', 'supplies'],
            'travel': ['נסיעות', 'דלק', 'travel'],
            'sales': ['מכירה', 'הכנסה', 'revenue', 'sale']
        }
        
        for category, words in keywords.items():
            if any(word in description for word in words):
                return category
        
        return 'other_expense' if tx.amount < 0 else 'other_income'
    
    def _get_default_budget(self) -> Dict[str, float]:
        """תקציב ברירת מחדל"""
        return {
            'sales': 500000,
            'services': 200000,
            'other_income': 50000,
            'salary': 180000,
            'rent': 25000,
            'utilities': 5000,
            'marketing': 30000,
            'software': 10000,
            'office': 5000,
            'travel': 8000,
            'insurance': 3000,
            'professional': 15000,
            'other_expense': 20000
        }
    
    def _generate_alert_message(self, cat: BudgetCategory) -> str:
        """יצירת הודעת התראה"""
        if cat.status == BudgetStatus.OVER_BUDGET:
            return f"חריגה בקטגוריית {cat.category_hebrew}: ₪{abs(cat.variance):,.0f} מעל התקציב ({abs(cat.variance_percentage):.1f}%)"
        elif cat.status == BudgetStatus.UNDER_BUDGET:
            return f"ביצוע נמוך בקטגוריית {cat.category_hebrew}: ₪{abs(cat.variance):,.0f} מתחת ליעד ({abs(cat.variance_percentage):.1f}%)"
        elif cat.status == BudgetStatus.WARNING:
            return f"אזהרה בקטגוריית {cat.category_hebrew}: מתקרבים לחריגה ({abs(cat.variance_percentage):.1f}%)"
        return ""
    
    def to_dict(self, obj) -> Dict:
        """המרה ל-dict"""
        if hasattr(obj, 'to_dict'):
            return obj.to_dict()
        return asdict(obj)
