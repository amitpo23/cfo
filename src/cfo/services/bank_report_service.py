"""
דוח מצב עסקי לבנק
Bank status report — composes real P&L, balance, cash position and AR/AP aging
into one banker-facing summary, with Excel export.
"""
from datetime import date, timedelta
from typing import Dict, Optional

from sqlalchemy import func
from sqlalchemy.orm import Session

from ..models import Bill, Organization
from .financial_reports_service import FinancialReportsService
from .financial_control_service import FinancialControlService
from .ar_service import AccountsReceivableService


class BankReportService:
    def __init__(self, db: Session, organization_id: int = 1):
        self.db = db
        self.organization_id = organization_id

    def generate(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> Dict:
        """הרכבת דוח מצב עסקי לבנק מנתונים אמיתיים."""
        today = end_date or date.today()
        period_start = start_date or today.replace(month=1, day=1)

        org = (
            self.db.query(Organization)
            .filter(Organization.id == self.organization_id)
            .first()
        )
        company_name = org.name if org else "העסק"

        reports = FinancialReportsService(self.db)
        pl = reports.generate_profit_loss(
            self.organization_id, period_start, today, compare_previous=False
        )
        # מקור אמת משותף עם מסך ה-KPI כדי שהמספרים יתאימו
        from .balance_snapshot import compute_balance_snapshot
        balance = compute_balance_snapshot(self.db, self.organization_id)

        control = FinancialControlService(
            self.db, organization_id=self.organization_id
        ).get_control_overview(period_start, today)

        aging = AccountsReceivableService(
            self.db, organization_id=self.organization_id
        ).get_aging_report(today)

        payables_total = float(
            self.db.query(func.coalesce(func.sum(Bill.balance), 0))
            .filter(
                Bill.organization_id == self.organization_id,
                Bill.balance > 0,
            ).scalar() or 0
        )

        return {
            "company_name": company_name,
            "generated_at": date.today().isoformat(),
            "period": {"start": period_start.isoformat(), "end": today.isoformat()},
            "profit_loss": {
                "total_revenue": float(pl.total_revenue or 0),
                "total_expenses": float(pl.total_expenses or 0),
                "gross_profit": float(pl.gross_profit or 0),
                "gross_margin": float(pl.gross_margin or 0),
                "operating_income": float(pl.operating_income or 0),
                "net_income": float(pl.net_income or 0),
                "net_margin": float(pl.net_margin or 0),
            },
            "balance_sheet": {
                "total_assets": balance["total_assets"],
                "total_current_assets": balance["current_assets"],
                "total_liabilities": balance["total_liabilities"],
                "total_current_liabilities": balance["current_liabilities"],
                "total_equity": balance["equity"],
                "current_ratio": balance["current_ratio"],
            },
            "cash_position": control.get("cash", {}),
            "receivables": {
                "total": float(aging.total_receivables or 0),
                "current": float(aging.current_total or 0),
                "days_31_60": float(aging.days_31_60_total or 0),
                "days_61_90": float(aging.days_61_90_total or 0),
                "days_91_120": float(aging.days_91_120_total or 0),
                "over_120": float(aging.over_120_total or 0),
            },
            "payables": {"total": payables_total},
        }

    def to_excel(self, report: Dict) -> bytes:
        """ייצוא דוח הבנק ל-Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ValueError("openpyxl required for Excel export")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "דוח מצב לבנק"
        ws.sheet_view.rightToLeft = True

        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12, color="FFFFFF")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        currency = "₪#,##0"
        percent = "0.0%"

        ws.merge_cells("A1:B1")
        ws["A1"] = f"דוח מצב עסקי — {report['company_name']}"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:B2")
        ws["A2"] = f"לתקופה {report['period']['start']} עד {report['period']['end']}"
        ws["A2"].alignment = Alignment(horizontal="center")

        row = 4

        def section(title: str):
            nonlocal row
            ws.merge_cells(f"A{row}:B{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font = section_font
            c.fill = section_fill
            row += 1

        def line(label: str, value, fmt: str = currency):
            nonlocal row
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = fmt
            row += 1

        pl = report["profit_loss"]
        section("רווח והפסד")
        line("סך הכנסות", pl["total_revenue"])
        line("סך הוצאות", pl["total_expenses"])
        line("רווח גולמי", pl["gross_profit"])
        line("שולי רווח גולמי", pl["gross_margin"], percent)
        line("רווח תפעולי", pl["operating_income"])
        line("רווח נקי", pl["net_income"])
        line("שולי רווח נקי", pl["net_margin"], percent)
        row += 1

        bs = report["balance_sheet"]
        section("מאזן")
        line("סך נכסים", bs["total_assets"])
        line("נכסים שוטפים", bs["total_current_assets"])
        line("סך התחייבויות", bs["total_liabilities"])
        line("התחייבויות שוטפות", bs["total_current_liabilities"])
        line("הון עצמי", bs["total_equity"])
        line("יחס שוטף", bs["current_ratio"], "0.00")
        row += 1

        cash = report["cash_position"]
        section("מצב מזומנים")
        line("יתרת בנק", cash.get("bank_account_balance", 0))
        line("תזרים נטו בתקופה", cash.get("net_flow", 0))
        row += 1

        ar = report["receivables"]
        section("חייבים (לקוחות)")
        line("סך חוב לקוחות", ar["total"])
        line("שוטף", ar["current"])
        line("31-60 יום", ar["days_31_60"])
        line("61-90 יום", ar["days_61_90"])
        line("91-120 יום", ar["days_91_120"])
        line("מעל 120 יום", ar["over_120"])
        row += 1

        section("זכאים (ספקים)")
        line("סך חוב לספקים", report["payables"]["total"])

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
