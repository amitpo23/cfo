"""
Financial Reports Service - דוחות כספיים מלאים
Profit & Loss, Balance Sheet, Cash Flow Projections
"""
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime, date, timedelta
from decimal import Decimal
from dataclasses import dataclass, asdict
from enum import Enum
import io
import logging

from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract

from ..models import (
    Transaction, Account, Organization, TransactionType, AccountType
)
from ..config import settings

logger = logging.getLogger(__name__)


class ReportPeriod(str, Enum):
    """תקופות דוח"""
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    CUSTOM = "custom"


@dataclass
class ProfitLossItem:
    """פריט בדוח רווח והפסד"""
    category: str
    category_hebrew: str
    amount: float
    percentage: float = 0.0
    previous_amount: float = 0.0
    change_percentage: float = 0.0


@dataclass
class ProfitLossReport:
    """דוח רווח והפסד"""
    period_start: str
    period_end: str
    revenue: List[ProfitLossItem]
    cost_of_goods_sold: List[ProfitLossItem]
    gross_profit: float
    gross_margin: float
    operating_expenses: List[ProfitLossItem]
    operating_income: float
    operating_margin: float
    other_income: List[ProfitLossItem]
    other_expenses: List[ProfitLossItem]
    net_income_before_tax: float
    tax_expense: float
    net_income: float
    net_margin: float
    total_revenue: float
    total_expenses: float
    
    def to_dict(self) -> Dict:
        return asdict(self)


@dataclass
class BalanceSheetItem:
    """פריט במאזן"""
    name: str
    name_hebrew: str
    amount: float
    previous_amount: float = 0.0


@dataclass
class BalanceSheetReport:
    """מאזן / דוח כספי"""
    as_of_date: str
    # נכסים
    current_assets: List[BalanceSheetItem]
    total_current_assets: float
    fixed_assets: List[BalanceSheetItem]
    total_fixed_assets: float
    other_assets: List[BalanceSheetItem]
    total_other_assets: float
    total_assets: float
    # התחייבויות
    current_liabilities: List[BalanceSheetItem]
    total_current_liabilities: float
    long_term_liabilities: List[BalanceSheetItem]
    total_long_term_liabilities: float
    total_liabilities: float
    # הון עצמי
    equity: List[BalanceSheetItem]
    total_equity: float
    # בדיקת איזון
    total_liabilities_and_equity: float
    is_balanced: bool
    
    def to_dict(self) -> Dict:
        return asdict(self)


@dataclass
class CashFlowProjectionItem:
    """פריט בתזרים חזוי"""
    month: str
    opening_balance: float
    inflows: float
    outflows: float
    net_flow: float
    closing_balance: float
    inflow_details: Dict[str, float]
    outflow_details: Dict[str, float]


@dataclass
class CashFlowProjectionReport:
    """דוח תזרים חזוי לבנק"""
    generated_date: str
    projection_months: int
    company_name: str
    prepared_by: str
    # נתונים היסטוריים
    historical_average_inflows: float
    historical_average_outflows: float
    # תחזית
    projections: List[CashFlowProjectionItem]
    # סיכום
    total_projected_inflows: float
    total_projected_outflows: float
    total_net_flow: float
    ending_balance: float
    minimum_balance: float
    maximum_balance: float
    # מדדים
    average_monthly_burn: float
    runway_months: float
    
    def to_dict(self) -> Dict:
        return asdict(self)


class FinancialReportsService:
    """
    שירות הפקת דוחות כספיים
    Financial Reports Generation Service
    """
    
    # מיפוי קטגוריות לעברית
    CATEGORY_HEBREW = {
        'sales': 'מכירות',
        'services': 'שירותים',
        'other_income': 'הכנסות אחרות',
        'interest_income': 'הכנסות ריבית',
        'rent_income': 'הכנסות שכירות',
        'cost_of_goods': 'עלות המכר',
        'materials': 'חומרים',
        'direct_labor': 'עבודה ישירה',
        'salaries': 'משכורות',
        'rent': 'שכירות',
        'utilities': 'חשמל ומים',
        'marketing': 'שיווק ופרסום',
        'insurance': 'ביטוח',
        'office_supplies': 'ציוד משרדי',
        'professional_fees': 'שירותים מקצועיים',
        'travel': 'נסיעות',
        'depreciation': 'פחת',
        'interest_expense': 'הוצאות ריבית',
        'bank_fees': 'עמלות בנק',
        'taxes': 'מסים',
        'other': 'אחר',
        'cash': 'מזומנים',
        'bank': 'בנק',
        'accounts_receivable': 'חייבים',
        'inventory': 'מלאי',
        'prepaid_expenses': 'הוצאות מראש',
        'equipment': 'ציוד',
        'vehicles': 'רכבים',
        'property': 'נדל"ן',
        'accounts_payable': 'ספקים',
        'short_term_loans': 'הלוואות לזמן קצר',
        'accrued_expenses': 'הוצאות לשלם',
        'long_term_loans': 'הלוואות לזמן ארוך',
        'mortgage': 'משכנתא',
        'share_capital': 'הון מניות',
        'retained_earnings': 'עודפים'
    }
    
    def __init__(self, db: Session):
        self.db = db
    
    def generate_profit_loss(
        self,
        organization_id: int,
        start_date: date,
        end_date: date,
        compare_previous: bool = True
    ) -> ProfitLossReport:
        """
        הפקת דוח רווח והפסד
        Generate Profit & Loss Statement
        """
        # שליפת כל העסקאות בתקופה
        transactions = self.db.query(Transaction).filter(
            Transaction.organization_id == organization_id,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date
        ).all()
        
        # חישוב תקופה קודמת להשוואה
        period_days = (end_date - start_date).days
        prev_end = start_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days)
        
        prev_transactions = []
        if compare_previous:
            prev_transactions = self.db.query(Transaction).filter(
                Transaction.organization_id == organization_id,
                Transaction.transaction_date >= prev_start,
                Transaction.transaction_date <= prev_end
            ).all()
        
        # ארגון נתונים לפי קטגוריה
        revenue_items = self._categorize_transactions(
            [t for t in transactions if t.transaction_type == TransactionType.INCOME],
            [t for t in prev_transactions if t.transaction_type == TransactionType.INCOME]
        )
        
        expense_items = self._categorize_transactions(
            [t for t in transactions if t.transaction_type == TransactionType.EXPENSE],
            [t for t in prev_transactions if t.transaction_type == TransactionType.EXPENSE]
        )
        
        # חישובים
        total_revenue = sum(item.amount for item in revenue_items)
        
        # הפרדת עלות מכר מהוצאות תפעוליות
        cogs_categories = ['cost_of_goods', 'materials', 'direct_labor']
        cogs_items = [i for i in expense_items if i.category in cogs_categories]
        operating_items = [i for i in expense_items if i.category not in cogs_categories]
        
        total_cogs = sum(item.amount for item in cogs_items)
        gross_profit = total_revenue - total_cogs
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue else 0
        
        total_operating = sum(item.amount for item in operating_items)
        operating_income = gross_profit - total_operating
        operating_margin = (operating_income / total_revenue * 100) if total_revenue else 0
        
        # הכנסות והוצאות אחרות
        other_income_items = [i for i in revenue_items if i.category in ['interest_income', 'other_income']]
        other_expense_items = [i for i in operating_items if i.category in ['interest_expense', 'other']]
        
        # רווח לפני מס
        total_other_income = sum(i.amount for i in other_income_items)
        total_other_expense = sum(i.amount for i in other_expense_items)
        net_before_tax = operating_income + total_other_income - total_other_expense
        
        # מס (הערכה 23%)
        tax_expense = max(0, net_before_tax * 0.23)
        net_income = net_before_tax - tax_expense
        net_margin = (net_income / total_revenue * 100) if total_revenue else 0
        
        # חישוב אחוזים מההכנסות
        for item in revenue_items:
            item.percentage = (item.amount / total_revenue * 100) if total_revenue else 0
        
        total_expenses = total_cogs + total_operating
        for item in expense_items:
            item.percentage = (item.amount / total_expenses * 100) if total_expenses else 0
        
        return ProfitLossReport(
            period_start=start_date.isoformat(),
            period_end=end_date.isoformat(),
            revenue=revenue_items,
            cost_of_goods_sold=cogs_items,
            gross_profit=float(gross_profit),
            gross_margin=float(gross_margin),
            operating_expenses=operating_items,
            operating_income=float(operating_income),
            operating_margin=float(operating_margin),
            other_income=other_income_items,
            other_expenses=other_expense_items,
            net_income_before_tax=float(net_before_tax),
            tax_expense=float(tax_expense),
            net_income=float(net_income),
            net_margin=float(net_margin),
            total_revenue=float(total_revenue),
            total_expenses=float(total_expenses)
        )
    
    def generate_balance_sheet(
        self,
        organization_id: int,
        as_of_date: date,
        compare_previous: bool = True
    ) -> BalanceSheetReport:
        """
        הפקת מאזן / דוח כספי
        Generate Balance Sheet
        """
        # שליפת חשבונות
        accounts = self.db.query(Account).filter(
            Account.organization_id == organization_id
        ).all()
        
        # יתרות עסקאות עד התאריך
        transactions = self.db.query(Transaction).filter(
            Transaction.organization_id == organization_id,
            Transaction.transaction_date <= as_of_date
        ).all()
        
        # חישוב יתרות לפי סוג חשבון
        balances = self._calculate_account_balances(accounts, transactions)
        
        # תקופה קודמת להשוואה
        prev_date = as_of_date - timedelta(days=365)
        prev_transactions = []
        if compare_previous:
            prev_transactions = self.db.query(Transaction).filter(
                Transaction.organization_id == organization_id,
                Transaction.transaction_date <= prev_date
            ).all()
        prev_balances = self._calculate_account_balances(accounts, prev_transactions)
        
        # נכסים שוטפים
        current_assets = [
            BalanceSheetItem('cash', 'מזומנים', balances.get('cash', 0), prev_balances.get('cash', 0)),
            BalanceSheetItem('bank', 'בנק', balances.get('bank', 0), prev_balances.get('bank', 0)),
            BalanceSheetItem('accounts_receivable', 'חייבים', balances.get('accounts_receivable', 0), prev_balances.get('accounts_receivable', 0)),
            BalanceSheetItem('inventory', 'מלאי', balances.get('inventory', 0), prev_balances.get('inventory', 0)),
            BalanceSheetItem('prepaid_expenses', 'הוצאות מראש', balances.get('prepaid_expenses', 0), prev_balances.get('prepaid_expenses', 0)),
        ]
        current_assets = [a for a in current_assets if a.amount != 0 or a.previous_amount != 0]
        total_current_assets = sum(a.amount for a in current_assets)
        
        # נכסים קבועים
        fixed_assets = [
            BalanceSheetItem('equipment', 'ציוד', balances.get('equipment', 0), prev_balances.get('equipment', 0)),
            BalanceSheetItem('vehicles', 'רכבים', balances.get('vehicles', 0), prev_balances.get('vehicles', 0)),
            BalanceSheetItem('property', 'נדל"ן', balances.get('property', 0), prev_balances.get('property', 0)),
        ]
        fixed_assets = [a for a in fixed_assets if a.amount != 0 or a.previous_amount != 0]
        total_fixed_assets = sum(a.amount for a in fixed_assets)
        
        # נכסים אחרים
        other_assets = [
            BalanceSheetItem('other_assets', 'נכסים אחרים', balances.get('other_assets', 0), prev_balances.get('other_assets', 0)),
        ]
        other_assets = [a for a in other_assets if a.amount != 0]
        total_other_assets = sum(a.amount for a in other_assets)
        
        total_assets = total_current_assets + total_fixed_assets + total_other_assets
        
        # התחייבויות שוטפות
        current_liabilities = [
            BalanceSheetItem('accounts_payable', 'ספקים', balances.get('accounts_payable', 0), prev_balances.get('accounts_payable', 0)),
            BalanceSheetItem('short_term_loans', 'הלוואות לז"ק', balances.get('short_term_loans', 0), prev_balances.get('short_term_loans', 0)),
            BalanceSheetItem('accrued_expenses', 'הוצאות לשלם', balances.get('accrued_expenses', 0), prev_balances.get('accrued_expenses', 0)),
        ]
        current_liabilities = [l for l in current_liabilities if l.amount != 0 or l.previous_amount != 0]
        total_current_liabilities = sum(l.amount for l in current_liabilities)
        
        # התחייבויות לזמן ארוך
        long_term_liabilities = [
            BalanceSheetItem('long_term_loans', 'הלוואות לז"א', balances.get('long_term_loans', 0), prev_balances.get('long_term_loans', 0)),
            BalanceSheetItem('mortgage', 'משכנתא', balances.get('mortgage', 0), prev_balances.get('mortgage', 0)),
        ]
        long_term_liabilities = [l for l in long_term_liabilities if l.amount != 0 or l.previous_amount != 0]
        total_long_term_liabilities = sum(l.amount for l in long_term_liabilities)
        
        total_liabilities = total_current_liabilities + total_long_term_liabilities
        
        # הון עצמי
        retained_earnings = total_assets - total_liabilities - balances.get('share_capital', 0)
        equity = [
            BalanceSheetItem('share_capital', 'הון מניות', balances.get('share_capital', 0), prev_balances.get('share_capital', 0)),
            BalanceSheetItem('retained_earnings', 'עודפים', retained_earnings, prev_balances.get('retained_earnings', 0)),
        ]
        total_equity = sum(e.amount for e in equity)
        
        total_liabilities_and_equity = total_liabilities + total_equity
        is_balanced = abs(total_assets - total_liabilities_and_equity) < 0.01
        
        return BalanceSheetReport(
            as_of_date=as_of_date.isoformat(),
            current_assets=current_assets,
            total_current_assets=float(total_current_assets),
            fixed_assets=fixed_assets,
            total_fixed_assets=float(total_fixed_assets),
            other_assets=other_assets,
            total_other_assets=float(total_other_assets),
            total_assets=float(total_assets),
            current_liabilities=current_liabilities,
            total_current_liabilities=float(total_current_liabilities),
            long_term_liabilities=long_term_liabilities,
            total_long_term_liabilities=float(total_long_term_liabilities),
            total_liabilities=float(total_liabilities),
            equity=equity,
            total_equity=float(total_equity),
            total_liabilities_and_equity=float(total_liabilities_and_equity),
            is_balanced=is_balanced
        )
    
    def generate_cash_flow_projection(
        self,
        organization_id: int,
        months: int = 12,
        opening_balance: Optional[float] = None
    ) -> CashFlowProjectionReport:
        """
        הפקת תזרים מזומנים חזוי לבנק
        Generate Projected Cash Flow for Bank
        """
        # שליפת נתונים היסטוריים (6 חודשים אחרונים)
        end_date = date.today()
        start_date = end_date - timedelta(days=180)
        
        transactions = self.db.query(Transaction).filter(
            Transaction.organization_id == organization_id,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date
        ).all()
        
        # חישוב ממוצעים היסטוריים
        income_txs = [t for t in transactions if t.transaction_type == TransactionType.INCOME]
        expense_txs = [t for t in transactions if t.transaction_type == TransactionType.EXPENSE]
        
        # קיבוץ לפי חודש
        monthly_income = self._group_by_month(income_txs)
        monthly_expense = self._group_by_month(expense_txs)
        
        # ממוצעים
        avg_income = sum(monthly_income.values()) / max(len(monthly_income), 1)
        avg_expense = sum(monthly_expense.values()) / max(len(monthly_expense), 1)
        
        # פירוט לפי קטגוריה
        income_by_category = self._sum_by_category(income_txs)
        expense_by_category = self._sum_by_category(expense_txs)
        
        # נרמול לחודש
        num_months = max(len(monthly_income), 1)
        monthly_income_detail = {k: v / num_months for k, v in income_by_category.items()}
        monthly_expense_detail = {k: v / num_months for k, v in expense_by_category.items()}
        
        # יתרת פתיחה
        if opening_balance is None:
            account = self.db.query(Account).filter(
                Account.organization_id == organization_id,
                Account.account_type == AccountType.ASSET
            ).first()
            opening_balance = float(account.balance) if account else 0
        
        # יצירת תחזית
        projections = []
        current_balance = opening_balance
        min_balance = current_balance
        max_balance = current_balance
        
        # גורמי עונתיות
        seasonality = self._calculate_seasonality(transactions)
        
        for i in range(months):
            proj_date = date.today() + timedelta(days=30 * (i + 1))
            month_name = proj_date.strftime('%Y-%m')
            month_num = proj_date.month
            
            # התאמת עונתיות
            season_factor = seasonality.get(month_num, 1.0)
            
            # חיזוי כניסות ויציאות
            projected_inflows = avg_income * season_factor
            projected_outflows = avg_expense
            
            # הוספת אי-ודאות קלה (±5%)
            import random
            random.seed(i)  # לשחזוריות
            projected_inflows *= (1 + random.uniform(-0.05, 0.05))
            projected_outflows *= (1 + random.uniform(-0.03, 0.03))
            
            net_flow = projected_inflows - projected_outflows
            closing = current_balance + net_flow
            
            # פירוט
            inflow_details = {
                self.CATEGORY_HEBREW.get(k, k): v * season_factor
                for k, v in monthly_income_detail.items()
            }
            outflow_details = {
                self.CATEGORY_HEBREW.get(k, k): v
                for k, v in monthly_expense_detail.items()
            }
            
            projections.append(CashFlowProjectionItem(
                month=month_name,
                opening_balance=round(current_balance, 2),
                inflows=round(projected_inflows, 2),
                outflows=round(projected_outflows, 2),
                net_flow=round(net_flow, 2),
                closing_balance=round(closing, 2),
                inflow_details={k: round(v, 2) for k, v in inflow_details.items()},
                outflow_details={k: round(v, 2) for k, v in outflow_details.items()}
            ))
            
            min_balance = min(min_balance, closing)
            max_balance = max(max_balance, closing)
            current_balance = closing
        
        # חישוב סיכומים
        total_inflows = sum(p.inflows for p in projections)
        total_outflows = sum(p.outflows for p in projections)
        total_net = total_inflows - total_outflows
        avg_burn = total_outflows / months
        
        # Runway
        if avg_burn > avg_income:
            runway = opening_balance / (avg_burn - avg_income)
        else:
            runway = float('inf')
        
        # קבלת שם החברה
        org = self.db.query(Organization).filter(
            Organization.id == organization_id
        ).first()
        company_name = org.name if org else "החברה"
        
        return CashFlowProjectionReport(
            generated_date=date.today().isoformat(),
            projection_months=months,
            company_name=company_name,
            prepared_by="CFO System",
            historical_average_inflows=round(avg_income, 2),
            historical_average_outflows=round(avg_expense, 2),
            projections=projections,
            total_projected_inflows=round(total_inflows, 2),
            total_projected_outflows=round(total_outflows, 2),
            total_net_flow=round(total_net, 2),
            ending_balance=round(current_balance, 2),
            minimum_balance=round(min_balance, 2),
            maximum_balance=round(max_balance, 2),
            average_monthly_burn=round(avg_burn, 2),
            runway_months=round(runway, 1) if runway != float('inf') else -1
        )
    
    def export_profit_loss_excel(
        self,
        report: ProfitLossReport
    ) -> bytes:
        """
        ייצוא דוח רווח והפסד ל-Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValueError("openpyxl required for Excel export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "דוח רווח והפסד"
        ws.sheet_view.rightToLeft = True
        
        # סגנונות
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        currency_format = '₪#,##0.00'
        percent_format = '0.0%'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        subtotal_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # כותרת
        ws.merge_cells('A1:D1')
        ws['A1'] = 'דוח רווח והפסד'
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f'לתקופה {report.period_start} עד {report.period_end}'
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # כותרות עמודות
        row = 4
        headers = ['פריט', 'סכום', 'אחוז', 'שינוי']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = title_font
            cell.fill = header_fill
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = thin_border
        
        row += 1
        
        # הכנסות
        ws.cell(row=row, column=1, value='הכנסות').font = title_font
        row += 1
        
        for item in report.revenue:
            ws.cell(row=row, column=1, value=item.category_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            ws.cell(row=row, column=3, value=item.percentage / 100).number_format = percent_format
            ws.cell(row=row, column=4, value=item.change_percentage / 100).number_format = percent_format
            row += 1
        
        # סה"כ הכנסות
        ws.cell(row=row, column=1, value='סה"כ הכנסות').font = title_font
        ws.cell(row=row, column=2, value=report.total_revenue).number_format = currency_format
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = subtotal_fill
        row += 2
        
        # עלות המכר
        ws.cell(row=row, column=1, value='עלות המכר').font = title_font
        row += 1
        
        for item in report.cost_of_goods_sold:
            ws.cell(row=row, column=1, value=item.category_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            row += 1
        
        # רווח גולמי
        ws.cell(row=row, column=1, value='רווח גולמי').font = title_font
        ws.cell(row=row, column=2, value=report.gross_profit).number_format = currency_format
        ws.cell(row=row, column=3, value=report.gross_margin / 100).number_format = percent_format
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = subtotal_fill
        row += 2
        
        # הוצאות תפעוליות
        ws.cell(row=row, column=1, value='הוצאות תפעוליות').font = title_font
        row += 1
        
        for item in report.operating_expenses:
            ws.cell(row=row, column=1, value=item.category_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            row += 1
        
        # רווח תפעולי
        ws.cell(row=row, column=1, value='רווח תפעולי').font = title_font
        ws.cell(row=row, column=2, value=report.operating_income).number_format = currency_format
        ws.cell(row=row, column=3, value=report.operating_margin / 100).number_format = percent_format
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = subtotal_fill
        row += 2
        
        # רווח נקי
        ws.cell(row=row, column=1, value='רווח לפני מס').font = title_font
        ws.cell(row=row, column=2, value=report.net_income_before_tax).number_format = currency_format
        row += 1
        
        ws.cell(row=row, column=1, value='מס')
        ws.cell(row=row, column=2, value=report.tax_expense).number_format = currency_format
        row += 1
        
        ws.cell(row=row, column=1, value='רווח נקי').font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value=report.net_income).number_format = currency_format
        ws.cell(row=row, column=2).font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value=report.net_margin / 100).number_format = percent_format
        
        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        
        # שמירה ל-bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_balance_sheet_excel(
        self,
        report: BalanceSheetReport
    ) -> bytes:
        """
        ייצוא מאזן ל-Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ValueError("openpyxl required for Excel export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "מאזן"
        ws.sheet_view.rightToLeft = True
        
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        currency_format = '₪#,##0.00'
        subtotal_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # כותרת
        ws.merge_cells('A1:C1')
        ws['A1'] = 'מאזן'
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:C2')
        ws['A2'] = f'ליום {report.as_of_date}'
        ws['A2'].alignment = Alignment(horizontal='center')
        
        row = 4
        
        # נכסים
        ws.cell(row=row, column=1, value='נכסים').font = header_font
        row += 1
        
        ws.cell(row=row, column=1, value='נכסים שוטפים').font = title_font
        row += 1
        
        for item in report.current_assets:
            ws.cell(row=row, column=1, value=item.name_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            ws.cell(row=row, column=3, value=item.previous_amount).number_format = currency_format
            row += 1
        
        ws.cell(row=row, column=1, value='סה"כ נכסים שוטפים').font = title_font
        ws.cell(row=row, column=2, value=report.total_current_assets).number_format = currency_format
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=2).fill = subtotal_fill
        row += 2
        
        ws.cell(row=row, column=1, value='נכסים קבועים').font = title_font
        row += 1
        
        for item in report.fixed_assets:
            ws.cell(row=row, column=1, value=item.name_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            row += 1
        
        ws.cell(row=row, column=1, value='סה"כ נכסים קבועים').font = title_font
        ws.cell(row=row, column=2, value=report.total_fixed_assets).number_format = currency_format
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=2).fill = subtotal_fill
        row += 2
        
        ws.cell(row=row, column=1, value='סה"כ נכסים').font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value=report.total_assets).number_format = currency_format
        ws.cell(row=row, column=2).font = Font(bold=True, size=14)
        row += 3
        
        # התחייבויות
        ws.cell(row=row, column=1, value='התחייבויות').font = header_font
        row += 1
        
        ws.cell(row=row, column=1, value='התחייבויות שוטפות').font = title_font
        row += 1
        
        for item in report.current_liabilities:
            ws.cell(row=row, column=1, value=item.name_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            row += 1
        
        ws.cell(row=row, column=1, value='סה"כ התחייבויות שוטפות').font = title_font
        ws.cell(row=row, column=2, value=report.total_current_liabilities).number_format = currency_format
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=2).fill = subtotal_fill
        row += 2
        
        ws.cell(row=row, column=1, value='סה"כ התחייבויות').font = title_font
        ws.cell(row=row, column=2, value=report.total_liabilities).number_format = currency_format
        row += 2
        
        # הון עצמי
        ws.cell(row=row, column=1, value='הון עצמי').font = header_font
        row += 1
        
        for item in report.equity:
            ws.cell(row=row, column=1, value=item.name_hebrew)
            ws.cell(row=row, column=2, value=item.amount).number_format = currency_format
            row += 1
        
        ws.cell(row=row, column=1, value='סה"כ הון עצמי').font = title_font
        ws.cell(row=row, column=2, value=report.total_equity).number_format = currency_format
        row += 2
        
        ws.cell(row=row, column=1, value='סה"כ התחייבויות והון').font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value=report.total_liabilities_and_equity).number_format = currency_format
        ws.cell(row=row, column=2).font = Font(bold=True, size=14)
        
        # התאמת רוחב
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_cash_flow_projection_excel(
        self,
        report: CashFlowProjectionReport
    ) -> bytes:
        """
        ייצוא תזרים חזוי ל-Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            raise ValueError("openpyxl required for Excel export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "תזרים חזוי"
        ws.sheet_view.rightToLeft = True
        
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=11)
        currency_format = '₪#,##0'
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # כותרת
        ws.merge_cells('A1:G1')
        ws['A1'] = f'תזרים מזומנים חזוי - {report.company_name}'
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f'תאריך הפקה: {report.generated_date}'
        ws['A3'] = f'תקופת תחזית: {report.projection_months} חודשים'
        
        # סיכום
        ws['A5'] = 'סיכום'
        ws['A5'].font = title_font
        
        ws['A6'] = 'ממוצע כניסות היסטורי:'
        ws['B6'] = report.historical_average_inflows
        ws['B6'].number_format = currency_format
        
        ws['A7'] = 'ממוצע יציאות היסטורי:'
        ws['B7'] = report.historical_average_outflows
        ws['B7'].number_format = currency_format
        
        ws['A8'] = 'יתרה צפויה בסוף התקופה:'
        ws['B8'] = report.ending_balance
        ws['B8'].number_format = currency_format
        
        ws['A9'] = 'יתרה מינימלית צפויה:'
        ws['B9'] = report.minimum_balance
        ws['B9'].number_format = currency_format
        if report.minimum_balance < 0:
            ws['B9'].fill = negative_fill
        
        ws['A10'] = 'Runway (חודשים):'
        ws['B10'] = report.runway_months if report.runway_months > 0 else '∞'
        
        # טבלת תחזית
        row = 13
        headers = ['חודש', 'יתרת פתיחה', 'כניסות', 'יציאות', 'תזרים נקי', 'יתרת סגירה']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        data_start_row = row
        
        for proj in report.projections:
            ws.cell(row=row, column=1, value=proj.month)
            ws.cell(row=row, column=2, value=proj.opening_balance).number_format = currency_format
            ws.cell(row=row, column=3, value=proj.inflows).number_format = currency_format
            ws.cell(row=row, column=4, value=proj.outflows).number_format = currency_format
            
            net_cell = ws.cell(row=row, column=5, value=proj.net_flow)
            net_cell.number_format = currency_format
            if proj.net_flow < 0:
                net_cell.fill = negative_fill
            else:
                net_cell.fill = positive_fill
            
            closing_cell = ws.cell(row=row, column=6, value=proj.closing_balance)
            closing_cell.number_format = currency_format
            if proj.closing_balance < 0:
                closing_cell.fill = negative_fill
            
            row += 1
        
        data_end_row = row - 1
        
        # סה"כ
        ws.cell(row=row, column=1, value='סה"כ').font = title_font
        ws.cell(row=row, column=3, value=report.total_projected_inflows).number_format = currency_format
        ws.cell(row=row, column=4, value=report.total_projected_outflows).number_format = currency_format
        ws.cell(row=row, column=5, value=report.total_net_flow).number_format = currency_format
        
        # גרף
        chart = LineChart()
        chart.title = "תחזית תזרים מזומנים"
        chart.style = 10
        chart.y_axis.title = "סכום (₪)"
        chart.x_axis.title = "חודש"
        
        # נתונים לגרף
        balance_data = Reference(ws, min_col=6, min_row=data_start_row - 1, max_row=data_end_row)
        months = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        
        chart.add_data(balance_data, titles_from_data=True)
        chart.set_categories(months)
        chart.width = 15
        chart.height = 8
        
        ws.add_chart(chart, f"H{data_start_row}")
        
        # התאמת רוחב עמודות
        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # ============= Helper Methods =============
    
    def _categorize_transactions(
        self,
        current: List[Transaction],
        previous: List[Transaction]
    ) -> List[ProfitLossItem]:
        """קיבוץ עסקאות לפי קטגוריה"""
        current_sums = {}
        for tx in current:
            cat = tx.category or 'other'
            current_sums[cat] = current_sums.get(cat, 0) + float(tx.amount)
        
        previous_sums = {}
        for tx in previous:
            cat = tx.category or 'other'
            previous_sums[cat] = previous_sums.get(cat, 0) + float(tx.amount)
        
        items = []
        all_cats = set(current_sums.keys()) | set(previous_sums.keys())
        
        for cat in all_cats:
            curr = current_sums.get(cat, 0)
            prev = previous_sums.get(cat, 0)
            change = ((curr - prev) / prev * 100) if prev else 0
            
            items.append(ProfitLossItem(
                category=cat,
                category_hebrew=self.CATEGORY_HEBREW.get(cat, cat),
                amount=curr,
                previous_amount=prev,
                change_percentage=change
            ))
        
        # מיון לפי סכום
        items.sort(key=lambda x: x.amount, reverse=True)
        return items
    
    def _calculate_account_balances(
        self,
        accounts: List[Account],
        transactions: List[Transaction]
    ) -> Dict[str, float]:
        """חישוב יתרות חשבונות"""
        balances = {}
        
        # יתרות מחשבונות
        for acc in accounts:
            key = acc.name.lower().replace(' ', '_')
            balances[key] = float(acc.balance)
        
        # התאמה לפי עסקאות
        for tx in transactions:
            if tx.transaction_type == TransactionType.INCOME:
                balances['bank'] = balances.get('bank', 0) + float(tx.amount)
            else:
                balances['bank'] = balances.get('bank', 0) - float(tx.amount)
        
        return balances
    
    def _group_by_month(self, transactions: List[Transaction]) -> Dict[str, float]:
        """קיבוץ עסקאות לפי חודש"""
        monthly = {}
        for tx in transactions:
            key = tx.transaction_date.strftime('%Y-%m')
            monthly[key] = monthly.get(key, 0) + float(tx.amount)
        return monthly
    
    def _sum_by_category(self, transactions: List[Transaction]) -> Dict[str, float]:
        """סיכום לפי קטגוריה"""
        sums = {}
        for tx in transactions:
            cat = tx.category or 'other'
            sums[cat] = sums.get(cat, 0) + float(tx.amount)
        return sums
    
    def _calculate_seasonality(self, transactions: List[Transaction]) -> Dict[int, float]:
        """חישוב עונתיות לפי חודש"""
        monthly_totals = {}
        for tx in transactions:
            if tx.transaction_type == TransactionType.INCOME:
                month = tx.transaction_date.month
                monthly_totals[month] = monthly_totals.get(month, [])
                monthly_totals[month].append(float(tx.amount))
        
        # ממוצע כללי
        all_values = [v for vals in monthly_totals.values() for v in vals]
        overall_avg = sum(all_values) / len(all_values) if all_values else 1
        
        # גורם עונתיות לכל חודש
        seasonality = {}
        for month in range(1, 13):
            if month in monthly_totals and monthly_totals[month]:
                month_avg = sum(monthly_totals[month]) / len(monthly_totals[month])
                seasonality[month] = month_avg / overall_avg if overall_avg else 1
            else:
                seasonality[month] = 1.0
        
        return seasonality
